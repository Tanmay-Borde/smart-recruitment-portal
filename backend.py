import os
from flask import Flask, flash, request, redirect, url_for, session, jsonify
import flask_cors
from werkzeug.utils import secure_filename
from flask_cors import CORS, cross_origin
import logging
import openpyxl
from openpyxl.styles import Font

# Create a new default style with a font size of 12
default_style = Font(size=12)

logging.basicConfig(level=logging.INFO)

logger = logging.getLogger(' - ')
logger.info("---------------------------------")

UPLOAD_FOLDER = 'C:/Users/tanmay.borde/Music'
ALLOWED_EXTENSIONS = set(['xlsx', 'pdf'])
FILE_EXTENSION = ''
FILE_NAME = ''

# JD Fields
JOB_DESCRIPTION = ''
JOB_POSITION = ''
PATH = ''

# End Result
PERCENTAGE = '97%'

app = Flask(__name__)
CORS(app)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


@app.route('/upload', methods=['POST'])
@cross_origin()
def fileUpload():
    global FILE_NAME
    global PATH
    target = os.path.join(UPLOAD_FOLDER, 'Veritas_JD_Folder')

    if not os.path.isdir(target):
        os.mkdir(target)
    file = request.files['file']
    filename = secure_filename(file.filename)
    PATH = "/".join([target, filename])
    FILE_NAME = filename

    # logger.info("PATH = %s", PATH)
    # logger.info("FILE_NAME = %s", FILE_NAME)

    file.save(PATH)
    session['uploadFilePath'] = PATH

    if os.path.isfile(PATH):
        if fileExtensionCheck() is False:
            return jsonify({"error": "Unexpected file extension!", }), 403
        else:
            if FILE_EXTENSION == "pdf":
                extractPDF()
            else:
                extractXLSX()
    else:
        return jsonify({"error": "File not uploaded, kindly refresh the page and try again!", }), 403

    return jsonify({"JOB_POSITION": JOB_POSITION, "PERCENTAGE": PERCENTAGE, "JOB_DESCRIPTION": JOB_DESCRIPTION}), 200


def fileExtensionCheck():
    global FILE_EXTENSION

    if "." in FILE_NAME:
        FILE_EXTENSION = FILE_NAME.split(".")[-1].lower()
        if FILE_EXTENSION == 'pdf':
            logger.info(" - PDF File Detected\n")
            FILE_EXTENSION = "pdf"
        if FILE_EXTENSION == '.xlsx':
            logger.info(" - xlsx File Detected\n")
            FILE_EXTENSION = "XLSX"
            return True
    else:
        return False


def extractPDF():
    logger.info(" - Extracting PDF\n")
    # Future Scope


def extractXLSX():
    global JOB_DESCRIPTION
    global JOB_POSITION
    logger.info(" - Extracting XLSX\n")
    workbook = openpyxl.load_workbook(PATH)
    workbook.default_style = default_style
    worksheet = workbook['Sheet1']
    JOB_DESCRIPTION = worksheet.cell(row=4, column=1).value
    JOB_POSITION = worksheet.cell(row=2, column=1).value


if __name__ == "__main__":
    app.secret_key = os.urandom(24)
    app.run(debug=True, use_reloader=False)  # host="0.0.0.0"

flask_cors.CORS(app, expose_headers='Authorization')
